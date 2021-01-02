from openpyxl import load_workbook

FILE_PATH = 'Planificacion DIINF diurno 2-2020.xlsx'
SHEET = 'Planificacion'

workbook = load_workbook(FILE_PATH,read_only= True)
sheet = workbook[SHEET]

class Ramo:
    tipoDeRamo = None
    
    def __init__(self,  codigoEjecu, codigoCivil, seccion, nivel, nombreRamo, nombreProfesor, apellidoProfesor, catedra, ejercicios, lab):
        if(codigoCivil == 'None'):
            self.codigoCivil = 'None'
        else:
            self.codigoCivil = str(codigoCivil)
        if(codigoEjecu == None):
            self.codigoEjecu = 'None'
        else:
            self.codigoEjecu = str(codigoEjecu)

        if(catedra == None):
            self.catedra = 'None'
        else:
            self.catedra = catedra
        if(ejercicios == None):
            self.ejercicios = 'None'
        else:
            self.ejercicios = ejercicios
        if(lab == None):
            self.lab = 'None'
        else:
            self.lab = lab
        
        self.seccion = seccion
        self.nivel = nivel
        self.nombreRamo = nombreRamo
        self.nombreProfesor = nombreProfesor
        self.apellidoProfesor = apellidoProfesor
        self.generarTipoDeRamo()
    
    def generarTipoDeRamo(self):
        aux = ''
        if(self.catedra != None and self.catedra != '' and self.catedra != 'None'):
            aux = aux + 'Catedra '
        if(self.ejercicios != None and self.ejercicios != '' and self.ejercicios != 'None'):
            aux = aux + 'Ejercicios '
        if(self.lab != None and self.lab != '' and self.lab != 'None'):
            aux = aux + 'Lab'
        self.tipoDeRamo = aux

    def toString(self):
        return self.codigoCivil  + ' ' + self.codigoEjecu  + ' ' + str(self.seccion)  + ' ' + str(self.nivel)  + ' ' + self.nombreRamo  + ' ' + self.nombreProfesor  + ' ' + self.apellidoProfesor  + ' ' + self.catedra  + ' ' + self.ejercicios + ' ' + self.lab  + ' ' + self.tipoDeRamo

def printAsignaturasPorNombre(asignaturas):
    array = []
    for ramo in asignaturas[0]:
        if(ramo.nombreRamo in array):
            continue
        else:
            array.append(ramo.nombreRamo)
    for i in array:
        print(i)

def printHorario(horario):
    for i in horario:
        print(i)

def getAsignaturas():
    ramos = []
    labs = []
    ayudantias = []

    for row in sheet.iter_rows(min_row=5, max_row=118, values_only=True):

        aux = Ramo(row[0], row[1], row[3], row[4], row[9], row[12], row[10], row[18], row[20], row[22])
        if(aux.ejercicios != 'None' and aux.catedra == 'None' and aux.lab == 'None'):
            ayudantias.append(aux)
        elif(aux.lab != 'None' and aux.catedra == 'None' and aux.ejercicios == 'None'):
            labs.append(aux)
        else:
            ramos.append(aux)
    asignaturas =[ramos, ayudantias, labs]
    
    return asignaturas

def getAsignaturasEjecu(nivel = 0):
    ramos = []
    labs = []
    ayudantias = []

    for row in sheet.iter_rows(min_row=5, max_row=118, values_only=True):
        if(row[0] == None or row[0] == 'None'):
            continue
        aux = Ramo(row[0], row[1], row[3], row[4], row[9], row[12], row[10], row[18], row[20], row[22])
        if(aux.ejercicios != 'None' and aux.catedra == 'None' and aux.lab == 'None'):
            ayudantias.append(aux)
        elif(aux.lab != 'None' and aux.catedra == 'None' and aux.ejercicios == 'None'):
            labs.append(aux)
        else:
            ramos.append(aux)
    asignaturas =[ramos, ayudantias, labs]
    
    return asignaturas

def getAsignaturasCivil(nivel = 0):
    ramos = []
    labs = []
    ayudantias = []

    for row in sheet.iter_rows(min_row=5, max_row=118, values_only=True):
        if(row[1] == None or row[1] == 'None'):
            continue
        aux = Ramo(row[0], row[1], row[3], row[4], row[9], row[12], row[10], row[18], row[20], row[22])
        if(aux.ejercicios != 'None' and aux.catedra == 'None' and aux.lab == 'None'):
            ayudantias.append(aux)
        elif(aux.lab != 'None' and aux.catedra == 'None' and aux.ejercicios == 'None'):
            labs.append(aux)
        else:
            ramos.append(aux)
    asignaturas =[ramos, ayudantias, labs]
    
    return asignaturas


def tieneLabAparte(codigo,labs):
    for ramo in labs:
        if((ramo.codigoCivil == codigo and ramo.codigoCivil != 'Null')):
            print(ramo.codigoCivil)
            return True
    return False

def tieneAyudantiaAparte(codigo,ayudantias):
    for ramo in ayudantias:
        if((ramo.codigoCivil == codigo and ramo.codigoCivil != 'Null') or (ramo.codigoEjecu == codigo and ramo.codigoEjecu != 'Null')):
            return True
    return False

def moduloToArray(modulo):
    diccionario = {
        'L': 0,
        'M': 1,
        'W': 2,
        'J': 3,
        'V': 4,
        'S': 5,
        '1': 0,
        '2': 1,
        '3': 2,
        '4': 3,
        '5': 4,
        '6': 5,
        '7': 6,
        '8': 7,
        '9': 8
    }
    if(modulo is None):
        return []
    else:
        return [diccionario[modulo[0]], diccionario[modulo[1]]]

def verificarAgregarRamo(horario, catedra = None, ayudantia = None, lab = None):
    condition = [False for x in range(3)]
    if(catedra != None and catedra != 'None'):
        condition[0] = True
    if(ayudantia != None and ayudantia != 'None'):
        condition[1] = True
    if(lab != None and lab != 'None'):
        condition[2] = True

    if(condition[0]):
        array = catedra.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            if(horario[aux[1]][aux[0]] != '|'):
                return False
            

    if(condition[1]):
        array = ayudantia.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            if(horario[aux[1]][aux[0]] != '|'):
                return False

    if(condition[2]):
        array = lab.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            if(horario[aux[1]][aux[0]] != '|'):
                return False
    return True

def agregaRamoHorario(horario, catedra = None, ayudantia = None, lab = None):
    condition = [False for x in range(3)]
    if(catedra != None and catedra != 'None'):
        condition[0] = True
    if(ayudantia != None and ayudantia != 'None'):
        condition[1] = True
    if(lab != None and lab != 'None'):
        condition[2] = True

    if(condition[0]):
        array = catedra.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            horario[aux[1]][aux[0]] = 'C'

    if(condition[1]):
        array = ayudantia.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            horario[aux[1]][aux[0]] = 'A'

    if(condition[2]):
        array = lab.split()
        for modulo in array:
            aux = moduloToArray(modulo)
            horario[aux[1]][aux[0]] = 'L'
    printHorario(horario)

def obtenerAsignaturas(opcion):
    if(opcion == 1):
        horario = [[None for x in range(6)] for y in range(9)]
        asignaturas = getAsignaturasCivil()
    elif(opcion == 2):
        horario = [[None for x in range(6)] for y in range(9)]
        asignaturas = getAsignaturasEjecu()
    else:
        exit()
    
def menu():
    opcion = -1
    while(opcion != 3):
        print('1. Ejecucion')
        print('2. Civil')
        print('3. Salir')
        opcion = int(input())
        if(opcion == 1):
            obtenerAsignaturas(opcion)
            exit()
        elif(opcion == 2):
            obtenerAsignaturas(opcion)
            exit()
#menu()


a = getAsignaturasCivil()
#print(a[0][0].nombreRamo)
#printAsignaturasPorNombre(a)
horario = [['|' for x in range(6)] for y in range(9)]


pepe = a[0][12]


agregaRamoHorario(horario, pepe.catedra, pepe.ejercicios, pepe.lab)
